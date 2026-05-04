# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Generate Whitebox Testing spreadsheet for Majunkita application."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
MODULE_FILL = PatternFill("solid", fgColor="E8F0FE")
MODULE_FONT = Font(bold=True, color="1A73E8", size=11)
NORMAL_FONT = Font(size=9)
BOLD_FONT = Font(bold=True, size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

HEADERS = ["No", "Fungsi / Method", "Teknik Pengujian", "Kasus Uji", "Input", "Expected Output", "Jalur/Cabang Kode", "Status"]
COL_WIDTHS = [6, 28, 18, 32, 35, 35, 40, 10]

# All test data per module
MODULES = {
    "1. Auth": {
        "file": "auth_repository.dart, auth_model.dart",
        "tests": [
            ["1.1","signIn()","Branch Coverage","Login dengan email (mengandung @)","identifier=\"user@mail.com\", password=\"pass123\"","AuthResponse sukses, langsung gunakan email","identifier.contains('@') → true","[ ]"],
            ["1.2","signIn()","Branch Coverage","Login dengan username (tanpa @)","identifier=\"admin\", password=\"pass123\"","Mencari email via _getEmailByUsername(), lalu login","identifier.contains('@') → false → _getEmailByUsername()","[ ]"],
            ["1.3","signIn()","Path Coverage","Login gagal — AuthException","identifier=\"user@mail.com\", password=\"wrong\"","throw Exception(_mapAuthError(e.message))","catch AuthException → _mapAuthError()","[ ]"],
            ["1.4","signIn()","Path Coverage","Login gagal — Exception umum","identifier=\"user@mail.com\" (network error)","throw Exception('Terjadi kesalahan saat login: ...')","catch (e) → generic error","[ ]"],
            ["1.5","_getEmailByUsername()","Branch Coverage","Username ditemukan di database","username=\"admin\"","Return email string","email != null && not empty → return email","[ ]"],
            ["1.6","_getEmailByUsername()","Branch Coverage","Username tidak ditemukan (null/empty)","username=\"tidakada\"","throw Exception('Username tidak ditemukan')","email == null || empty → throw","[ ]"],
            ["1.7","_getEmailByUsername()","Path Coverage","Error lain saat RPC call","username=\"admin\" (DB timeout)","throw Exception('Username tidak ditemukan')","catch → !contains('Username tidak ditemukan') → throw generic","[ ]"],
            ["1.8","_mapAuthError()","Branch Coverage","Error: invalid login credentials","message=\"Invalid login credentials\"","\"Email atau Password salah.\"","contains('invalid login credentials') → true","[ ]"],
            ["1.9","_mapAuthError()","Branch Coverage","Error: email not confirmed","message=\"Email not confirmed\"","\"Email belum diverifikasi. Silakan cek inbox Anda.\"","contains('email not confirmed') → true","[ ]"],
            ["1.10","_mapAuthError()","Branch Coverage","Error tidak dikenali","message=\"Some unknown error\"","\"Some unknown error\" (pesan asli)","no match → return original message","[ ]"],
            ["1.11","getCurrentUserProfile()","Branch Coverage","User null (belum login)","currentUser == null","return null","user == null → return null","[ ]"],
            ["1.12","getCurrentUserProfile()","Path Coverage","User ada, profil ditemukan","currentUser != null, profile exists","return Map profil","user != null → query → return data","[ ]"],
            ["1.13","getCurrentUserProfile()","Path Coverage","User ada, profil tidak ada","currentUser != null, no profile","return null","user != null → catch (e) → return null","[ ]"],
            ["1.14","createUserByAdmin()","Branch Coverage","Berhasil membuat user baru","email, password, name, role, noTelp valid","void (sukses)","data != Map || !containsKey('error') → success","[ ]"],
            ["1.15","createUserByAdmin()","Branch Coverage","Edge Function mengembalikan error","email duplikat","throw Exception(data['error'])","data is Map && containsKey('error') → throw","[ ]"],
            ["1.16","createUserByAdmin()","Path Coverage","FunctionException (HTTP error)","Forbidden access","throw Exception(message)","catch FunctionException → details/reasonPhrase","[ ]"],
            ["1.17","Profiles.fromMap()","Statement Coverage","Parsing data profil lengkap","Map dengan semua field","Profiles object dengan semua field terisi","All fields mapped correctly","[ ]"],
            ["1.18","Profiles.toMap()","Statement Coverage","Konversi Profiles ke Map","Profiles object valid","Map dengan key id, name, username, email, role, no_telp","All fields serialized","[ ]"],
        ]
    },
    "2. Perca": {
        "file": "perca_repository.dart, perca_stock_model.dart",
        "tests": [
            ["2.1","generateSackCode()","Branch Coverage","Perca jenis Kaos, berat bulat","percaType=\"kaos\", weight=45.0","\"K-45\"","toLowerCase()=='kaos' → 'K', weight==roundToDouble → toInt","[ ]"],
            ["2.2","generateSackCode()","Branch Coverage","Perca jenis Kain, berat bulat","percaType=\"kain\", weight=30.0","\"B-30\"","toLowerCase()!='kaos' → 'B', weight==roundToDouble → toInt","[ ]"],
            ["2.3","generateSackCode()","Branch Coverage","Perca jenis Kaos, berat desimal","percaType=\"kaos\", weight=25.5","\"K-25.50\"","prefix='K', weight!=roundToDouble → toStringAsFixed(2)","[ ]"],
            ["2.4","PercasStock.toJson()","Statement Coverage","Serialisasi ke JSON","PercasStock object valid","Map: id_factory, date_entry, perca_type, weight, delivery_proof, sack_code","All fields mapped","[ ]"],
            ["2.5","uploadImageToStorage()","Branch Coverage","Upload berhasil","File gambar valid","Return public URL string","response.isNotEmpty → getPublicUrl","[ ]"],
            ["2.6","uploadImageToStorage()","Branch Coverage","Upload gagal — response kosong","File valid, storage error","throw Exception(\"Failed to upload image...\")","response.isEmpty → throw","[ ]"],
            ["2.7","uploadImageToStorage()","Path Coverage","Upload gagal — exception","File corrupt / network error","throw Exception('Gagal upload gambar: ...')","catch (e) → throw","[ ]"],
            ["2.8","saveStockToDatabase()","Statement Coverage","Simpan stok berhasil","PercasStock object valid","void (sukses insert)","insert → success","[ ]"],
            ["2.9","saveMultipleStocksToDatabase()","Statement Coverage","Simpan multiple stocks berhasil","List<PercasStock> berisi 3 item","void (sukses batch insert)","map toJson → insert list","[ ]"],
            ["2.10","getPercaHistory()","Statement Coverage","Ambil riwayat berhasil","—","List<Map> terurut date_entry DESC","select + order → return list","[ ]"],
        ]
    },
    "3. Tailor": {
        "file": "tailor_repository.dart, tailor_model.dart, prediction_model_utils.dart",
        "tests": [
            ["3.1","TailorModel.fromJson()","Statement Coverage","Parsing JSON lengkap","JSON dengan semua field","TailorModel dengan field terisi benar","all safe null handling paths","[ ]"],
            ["3.2","TailorModel.fromJson()","Branch Coverage","Parsing JSON dengan null values","JSON: {id:null, name:null, ...}","TailorModel dengan default values ('', 0, DateTime.now())","null coalescing ?? paths","[ ]"],
            ["3.3","searchTailors()","Branch Coverage","Query kosong","query=\"\"","Return getAllTailors()","query.isEmpty → true → getAllTailors()","[ ]"],
            ["3.4","searchTailors()","Branch Coverage","Query tidak kosong","query=\"Budi\"","Return filtered tailors matching \"Budi\"","query.isEmpty → false → ilike query","[ ]"],
            ["3.5","getTailorById()","Branch Coverage","Tailor ditemukan","id=\"uuid-valid\"","TailorModel object","response != null → return TailorModel","[ ]"],
            ["3.6","getTailorById()","Branch Coverage","Tailor tidak ditemukan","id=\"uuid-invalid\"","return null","response == null → return null","[ ]"],
            ["3.7","createTailor()","Path Coverage","Berhasil membuat tailor","name, noTelp, address valid","TailorModel baru","insert → select → single → success","[ ]"],
            ["3.8","createTailor()","Branch Coverage","Duplicate key error","name yang sudah ada","throw 'Data penjahit sudah terdaftar...'","catch → contains('duplicate key') → throw user-friendly","[ ]"],
            ["3.9","createTailor()","Path Coverage","Error umum","Network error","throw 'Gagal membuat data penjahit: ...'","catch → !duplicate → throw generic","[ ]"],
            ["3.10","updateTailor()","Branch Coverage","Update dengan gambar baru (hapus lama)","tailorImages=\"new.jpg\", oldImageUrl=\"old.jpg\"","TailorModel updated, gambar lama dihapus","new!=null && old!=null && new!=old → deleteTailorImage","[ ]"],
            ["3.11","updateTailor()","Branch Coverage","Update tanpa gambar baru","tailorImages=null","TailorModel updated, gambar lama tetap","tailorImages==null → skip delete","[ ]"],
            ["3.12","updateTailor()","Path Coverage","Hapus gambar lama gagal tapi update jalan","tailorImages=\"new.jpg\" (storage error)","TailorModel updated (warning logged, no throw)","deleteTailorImage catch → continue update","[ ]"],
            ["3.13","deleteTailor()","Branch Coverage","Hapus tailor dengan gambar","id valid, tailorImages != null","Tailor + gambar dihapus","tailor!=null && tailorImages!=null → delete image+folder","[ ]"],
            ["3.14","deleteTailor()","Branch Coverage","Hapus tailor tanpa gambar","id valid, tailorImages == null","Tailor dihapus, skip image delete","tailorImages==null → skip deleteTailorImage","[ ]"],
            ["3.15","getTailorEfficiencyStats()","Branch Coverage","RPC mengembalikan data valid","tailorId valid, ada data","Map dengan reff, prediksi_majun, dll.","rows.isNotEmpty → parse all _d() fields","[ ]"],
            ["3.16","getTailorEfficiencyStats()","Branch Coverage","RPC mengembalikan kosong","tailorId baru (belum ada data)","Map dengan semua value = 0","rows.isEmpty → return zeros map","[ ]"],
            ["3.17","calculateReff()","Branch Coverage","TotalPercaDiambil > 0","totalMajunDisetor=80, totalPercaDiambil=100","0.8","totalPercaDiambil > 0 → division","[ ]"],
            ["3.18","calculateReff()","Branch Coverage","TotalPercaDiambil = 0 (penjahit baru)","totalMajunDisetor=0, totalPercaDiambil=0","0.0","totalPercaDiambil <= 0 → return 0.0","[ ]"],
            ["3.19","calculatePrediksiMajun()","Statement Coverage","Prediksi dengan data valid","sisaPerca=50, reff=0.8","40.0","sisaPerca * reff = 40.0","[ ]"],
            ["3.20","calculatePrediksiMajun()","Boundary Value","Prediksi dengan reff = 0","sisaPerca=50, reff=0","0.0","50 * 0 = 0.0","[ ]"],
            ["3.21","calculateEfficiencyStats()","Path Coverage","Hitung statistik lengkap","totalMajun=80, totalPerca=100, sisaPerca=50","{reff: 0.8, prediksi_majun: 40.0}","calculateReff → calculatePrediksiMajun → return map","[ ]"],
            ["3.22","createSalaryWithdrawal()","Branch Coverage","Penarikan upah berhasil","tailorId, amount=500000, dateEntry valid","SalaryWithdrawalModel + balance updated","insert → getTailorById != null → update balance","[ ]"],
            ["3.23","createSalaryWithdrawal()","Branch Coverage","Tailor null (edge case)","tailorId invalid","SalaryWithdrawalModel (balance not updated)","insert → getTailorById == null → skip update","[ ]"],
        ]
    },
    "4. Factory": {
        "file": "factory_repository.dart, factory_model.dart",
        "tests": [
            ["4.1","getAllFactories()","Statement Coverage","Ambil data factory dengan pagination","page=1, limit=50","List<FactoryModel>","offset=(1-1)*50=0, range(0,49)","[ ]"],
            ["4.2","searchFactories()","Branch Coverage","Query kosong","query=\"\"","Return getAllFactories()","query.isEmpty → true","[ ]"],
            ["4.3","searchFactories()","Branch Coverage","Query valid","query=\"Pabrik A\"","List filtered factories","query.isEmpty → false → ilike","[ ]"],
            ["4.4","createFactory()","Branch Coverage","Berhasil membuat factory","factoryName, address, noTelp valid","FactoryModel baru","insert → success","[ ]"],
            ["4.5","createFactory()","Branch Coverage","Duplicate key error","factoryName yang sudah ada","throw user-friendly error","catch → contains('duplicate') → throw","[ ]"],
            ["4.6","deleteFactory()","Branch Coverage","Hapus factory yang masih ada relasi","id dengan foreign key constraint","throw 'Tidak bisa menghapus factory...'","catch → contains('violates foreign key') → throw","[ ]"],
            ["4.7","deleteFactory()","Path Coverage","Hapus factory berhasil","id tanpa relasi","void (sukses)","delete → success","[ ]"],
        ]
    },
    "5. Majun": {
        "file": "majun_repository.dart, majun_transactions_model.dart",
        "tests": [
            ["5.1","getMajunPricePerKg()","Statement Coverage","Ambil harga per kg berhasil","—","double (e.g. 5000.0)","select → parse → return double","[ ]"],
            ["5.2","getMajunPricePerKg()","Branch Coverage","Value null di database","app_settings value = null","0.0","double.tryParse(null → '0') → 0.0","[ ]"],
            ["5.3","setorMajun()","Statement Coverage","Setor majun berhasil","tailorId, weightMajun=10.5, deliveryProof, staffId","SetorMajunResult dengan earnedWage","insert → select → single → fromJson","[ ]"],
            ["5.4","setorLimbah()","Statement Coverage","Setor limbah berhasil","tailorId, weightLimbah=5.0, staffId","LimbahTransactionsModel","insert → select → single → fromJson","[ ]"],
            ["5.5","getMajunHistory()","Branch Coverage","Response berupa List","limit=50, offset=0","List<MajunTransactionsModel>","response is List → true → map fromJson","[ ]"],
            ["5.6","getMajunHistory()","Branch Coverage","Response bukan List","limit=50, offset=0 (RPC null)","[] (empty list)","response is List → false → return []","[ ]"],
            ["5.7","getMonthlyMajunStats()","Branch Coverage","Data bulanan ada","—","Map<String,double> e.g. {\"2026-04\": 150.0}","response is List → monthKey.isNotEmpty","[ ]"],
            ["5.8","getMonthlyMajunStats()","Branch Coverage","MonthKey kosong (data corrupt)","item['month_key'] = null","Skip item, tidak masuk Map","monthKey.isEmpty → skip","[ ]"],
            ["5.9","MajunTransactionsModel.fromJson()","Statement Coverage","Parsing JSON lengkap","JSON dengan semua field","MajunTransactionsModel valid","all safe parse paths","[ ]"],
            ["5.10","SetorMajunResult.fromJson()","Statement Coverage","Parsing JSON result","JSON: {id, weight_majun, earned_wage}","SetorMajunResult object","all double.tryParse paths","[ ]"],
        ]
    },
    "6. Expedisi": {
        "file": "expedition_repository.dart, expedition_model.dart",
        "tests": [
            ["6.1","createExpedition()","Path Coverage","Berhasil buat expedisi + upload foto","ExpeditionModel valid, File gambar","void (insert + upload sukses)","compress → upload → insert → success","[ ]"],
            ["6.2","createExpedition()","Branch Coverage","Upload gagal — response kosong","File valid, storage error","throw Exception","response.isEmpty → throw","[ ]"],
            ["6.3","deleteExpedition()","Path Coverage","Hapus expedisi + file storage berhasil","id, imageUrl valid","void (delete DB + storage)","delete DB → extract path → remove storage","[ ]"],
            ["6.4","getAvailableStock()","Statement Coverage","Hitung stok gudang tersedia","—","double (SUM majun - SUM expeditions)","rpc → return value","[ ]"],
            ["6.5","getWeightPerSack()","Branch Coverage","Setting ditemukan","—","int (e.g. 25)","select → parse → return int","[ ]"],
            ["6.6","updateWeightPerSack()","Statement Coverage","Update setting berhasil","newWeight=30","void (update sukses)","update → success","[ ]"],
        ]
    },
    "7. Partner": {
        "file": "manage_partner_repository.dart, manage_partner_models.dart",
        "tests": [
            ["7.1","_getUsersByRole()","Branch Coverage","Fetch tanpa search query","role='admin', searchQuery=null","List<Admin> semua admin","searchQuery==null → skip ilike → order → return","[ ]"],
            ["7.2","_getUsersByRole()","Branch Coverage","Fetch dengan search query","role='admin', searchQuery='Budi'","List<Admin> filtered","searchQuery!=null && isNotEmpty → ilike → order → return","[ ]"],
            ["7.3","_getUsersByRole()","Branch Coverage","Search query kosong","role='admin', searchQuery=''","List<Admin> semua admin (skip ilike)","searchQuery.isEmpty → skip ilike","[ ]"],
            ["7.4","_getUsersByRole()","Path Coverage","Error saat fetch","role='admin' (DB error)","throw Exception('Gagal mengambil data admin: ...')","catch (e) → throw","[ ]"],
            ["7.5","_createUserViaFunction()","Branch Coverage","Berhasil buat user (status 200)","email, password, name, role valid","void (sukses)","response.status == 200 → success","[ ]"],
            ["7.6","_createUserViaFunction()","Branch Coverage","Gagal — status != 200, ada error detail","email duplikat","throw Exception('error (details)')","status != 200 → data is Map → error+details != null","[ ]"],
            ["7.7","_createUserViaFunction()","Branch Coverage","Gagal — status != 200, tanpa error detail","server error","throw Exception('Gagal: Status 500')","status != 200 → data bukan Map → throw status","[ ]"],
            ["7.8","_createUserViaFunction()","Path Coverage","Exception umum","network error","throw Exception('Gagal membuat user: ...')","catch (e) → throw","[ ]"],
            ["7.9","createAdmin()","Statement Coverage","Berhasil membuat admin baru","name, email, noTelp, password valid","void (admin dibuat via Edge Function)","email.split('@')[0] → username → _createUserViaFunction(role:admin)","[ ]"],
            ["7.10","createDriver()","Statement Coverage","Berhasil membuat driver baru","name, email, noTelp, password valid","void (driver dibuat via Edge Function)","email.split('@')[0] → username → _createUserViaFunction(role:driver)","[ ]"],
            ["7.11","updateUser()","Branch Coverage","Update dengan password baru","id, name, email, noTelp, password='new123'","void (user updated)","password!=null && isNotEmpty → body['password'] = password","[ ]"],
            ["7.12","updateUser()","Branch Coverage","Update tanpa password","id, name, email, noTelp, password=null","void (user updated tanpa ubah password)","password==null → skip password field","[ ]"],
            ["7.13","updateUser()","Branch Coverage","Update dengan username","id, name, email, noTelp, username='admin1'","void (user updated + username)","username!=null && isNotEmpty → body['username']","[ ]"],
            ["7.14","updateUser()","Branch Coverage","Update dengan role dan address","id, name, role='driver', address='Jl. X'","void (role + address updated)","role!=null → body['role'], address!=null → body['address']","[ ]"],
            ["7.15","updateUser()","Path Coverage","Update gagal — status != 200","id valid, server error","throw Exception('Gagal mengupdate user: Status ...')","response.status != 200 → throw","[ ]"],
            ["7.16","deleteUser()","Path Coverage","Hapus user berhasil","id valid","void (user dihapus via Edge Function)","invoke delete-user → status 200 → success","[ ]"],
            ["7.17","deleteUser()","Path Coverage","Hapus user gagal — status != 200","id valid, server error","throw Exception('Gagal menghapus user: Status ...')","response.status != 200 → throw","[ ]"],
            ["7.18","Admin.fromJson()","Statement Coverage","Parsing JSON admin lengkap","JSON dengan semua field","Admin object","all fields mapped, name fallback 'Tanpa Nama'","[ ]"],
            ["7.19","Driver.fromJson()","Statement Coverage","Parsing JSON driver lengkap","JSON dengan semua field","Driver object","all fields mapped, name fallback 'Tanpa Nama'","[ ]"],
            ["7.20","UserProfile.fromJson()","Branch Coverage","Parsing JSON dengan null values","JSON: {name:null, email:null}","UserProfile: name='Tanpa Nama', email='', noTelp='-', role='staff'","null coalescing ?? fallback values","[ ]"],
        ]
    },
    "8. Core Utilities": {
        "file": "currency_helper.dart, image_compressor.dart, storage_service.dart",
        "tests": [
            ["8.1","CurrencyHelper.formatRupiah()","Statement Coverage","Format angka ke Rupiah","value=100000","\"Rp100.000\"","currencyFormat.format()","[ ]"],
            ["8.2","CurrencyHelper.formatRupiah()","Boundary Value","Format angka 0","value=0","\"Rp0\"","format(0)","[ ]"],
            ["8.3","ImageCompressor.compressImage()","Branch Coverage","Kompresi berhasil","File gambar valid (> 1024px)","File terkompresi (lebih kecil)","result != null → return File(result.path)","[ ]"],
            ["8.4","ImageCompressor.compressImage()","Branch Coverage","Kompresi return null","File valid, compress return null","Return file asli (fallback)","result == null → return file","[ ]"],
            ["8.5","ImageCompressor.compressImage()","Path Coverage","Kompresi gagal — exception","File corrupt","Return file asli (fallback)","catch (_) → return file","[ ]"],
            ["8.6","StorageService.deleteTailorImage()","Branch Coverage","URL valid — hapus berhasil","imageUrl valid (contains 'majunkita')","void (file dihapus)","bucketIndex != -1 → extract path → remove","[ ]"],
            ["8.7","StorageService.deleteTailorImage()","Branch Coverage","URL format salah","imageUrl tanpa 'majunkita'","Warning logged, tidak throw","bucketIndex == -1 → throw → catch → log warn","[ ]"],
            ["8.8","deleteTailorImageFolder()","Branch Coverage","Folder kosong","tailorId (no files)","Return early","files.isEmpty → return","[ ]"],
            ["8.9","deleteTailorImageFolder()","Branch Coverage","Tidak ada file cocok pattern","tailorId tidak cocok","Return early","tailorFiles.isEmpty → return","[ ]"],
            ["8.10","deleteTailorImageFolder()","Path Coverage","File ditemukan — hapus batch","tailorId cocok 3 files","3 files dihapus","tailorFiles.isNotEmpty → map → remove batch","[ ]"],
        ]
    },
}

def create_sheet(ws, module_name, file_info, tests):
    # Title row
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"Modul {module_name} — {file_info}"
    cell.font = MODULE_FONT
    cell.fill = MODULE_FILL
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, test in enumerate(tests, 3):
        for col_idx, value in enumerate(test, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx == 1:  # No
                cell.alignment = CENTER
            elif col_idx == 8:  # Status
                cell.alignment = CENTER
                cell.font = BOLD_FONT
            else:
                cell.alignment = WRAP

    # Freeze panes
    ws.freeze_panes = "A3"

# Create sheets
first = True
for module_name, data in MODULES.items():
    if first:
        ws = wb.active
        ws.title = module_name
        first = False
    else:
        ws = wb.create_sheet(title=module_name)
    create_sheet(ws, module_name, data["file"], data["tests"])

# Summary sheet
ws_summary = wb.create_sheet(title="Ringkasan", index=0)
ws_summary.merge_cells("A1:D1")
cell = ws_summary["A1"]
cell.value = "Ringkasan Whitebox Testing — Aplikasi Majunkita"
cell.font = Font(bold=True, size=14, color="1A73E8")
cell.alignment = Alignment(vertical="center")

ws_summary.merge_cells("A2:D2")
cell = ws_summary["A2"]
cell.value = "Tanggal: 23 April 2026 | Flutter + Supabase"
cell.font = Font(size=10, color="666666")

summary_headers = ["No", "Modul", "File Sumber", "Jumlah Kasus Uji"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=4, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_summary.column_dimensions["A"].width = 6
ws_summary.column_dimensions["B"].width = 28
ws_summary.column_dimensions["C"].width = 50
ws_summary.column_dimensions["D"].width = 20

total = 0
for row_idx, (module_name, data) in enumerate(MODULES.items(), 5):
    count = len(data["tests"])
    total += count
    ws_summary.cell(row=row_idx, column=1, value=row_idx - 4).font = NORMAL_FONT
    ws_summary.cell(row=row_idx, column=1).alignment = CENTER
    ws_summary.cell(row=row_idx, column=1).border = THIN_BORDER
    ws_summary.cell(row=row_idx, column=2, value=module_name).font = NORMAL_FONT
    ws_summary.cell(row=row_idx, column=2).border = THIN_BORDER
    ws_summary.cell(row=row_idx, column=3, value=data["file"]).font = NORMAL_FONT
    ws_summary.cell(row=row_idx, column=3).border = THIN_BORDER
    ws_summary.cell(row=row_idx, column=4, value=count).font = NORMAL_FONT
    ws_summary.cell(row=row_idx, column=4).alignment = CENTER
    ws_summary.cell(row=row_idx, column=4).border = THIN_BORDER

total_row = 5 + len(MODULES)
ws_summary.cell(row=total_row, column=3, value="TOTAL").font = BOLD_FONT
ws_summary.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
ws_summary.cell(row=total_row, column=3).border = THIN_BORDER
ws_summary.cell(row=total_row, column=4, value=total).font = Font(bold=True, size=11, color="1A73E8")
ws_summary.cell(row=total_row, column=4).alignment = CENTER
ws_summary.cell(row=total_row, column=4).border = THIN_BORDER

ws_summary.freeze_panes = "A5"

output = "/home/dwirez/PROJEK/T_A/majunkita/docs/whitebox_testing.xlsx"
wb.save(output)
print(f"✅ Saved: {output}")
print(f"📊 Total test cases: {total} across {len(MODULES)} modules")
