# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Generate Blackbox Testing spreadsheet for Majunkita application."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="0D8043")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
MODULE_FILL = PatternFill("solid", fgColor="E6F4EA")
MODULE_FONT = Font(bold=True, color="0D8043", size=11)
NORMAL_FONT = Font(size=9)
BOLD_FONT = Font(bold=True, size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

HEADERS = ["No", "Skenario Pengujian", "Langkah-Langkah", "Data Input", "Expected Result", "Actual Result", "Status (Pass/Fail)"]
COL_WIDTHS = [6, 30, 42, 30, 32, 25, 14]

MODULES = {
    "1. Login": {
        "desc": "Halaman Login — login_screen.dart",
        "tests": [
            ["1.1", "Login dengan email dan password valid", "1. Buka aplikasi\n2. Isi field 'Email atau Username' dengan email\n3. Isi field 'Password'\n4. Klik tombol 'MASUK'", "identifier: admin@mail.com\npassword: password123", "Muncul SnackBar 'Login berhasil!' dan diarahkan ke Dashboard sesuai role", "", "[ ]"],
            ["1.2", "Login dengan username dan password valid", "1. Buka aplikasi\n2. Isi field 'Email atau Username' dengan username\n3. Isi field 'Password'\n4. Klik tombol 'MASUK'", "identifier: admin\npassword: password123", "Muncul SnackBar 'Login berhasil!' dan diarahkan ke Dashboard sesuai role", "", "[ ]"],
            ["1.3", "Login dengan password salah", "1. Buka aplikasi\n2. Isi email/username yang benar\n3. Isi password yang salah\n4. Klik tombol 'MASUK'", "identifier: admin@mail.com\npassword: wrongpass", "Muncul SnackBar merah 'Email atau Password salah.'", "", "[ ]"],
            ["1.4", "Login dengan username tidak terdaftar", "1. Buka aplikasi\n2. Isi username yang tidak ada di database\n3. Isi password\n4. Klik tombol 'MASUK'", "identifier: usertidakada\npassword: pass123", "Muncul SnackBar merah 'Username tidak ditemukan'", "", "[ ]"],
            ["1.5", "Login dengan field kosong", "1. Buka aplikasi\n2. Biarkan field email/username kosong\n3. Biarkan field password kosong\n4. Klik tombol 'MASUK'", "identifier: (kosong)\npassword: (kosong)", "Muncul pesan validasi 'Email atau Username wajib diisi' dan 'Password wajib diisi'", "", "[ ]"],
            ["1.6", "Login dengan email kosong saja", "1. Buka aplikasi\n2. Biarkan field email kosong\n3. Isi password\n4. Klik tombol 'MASUK'", "identifier: (kosong)\npassword: password123", "Muncul pesan validasi 'Email atau Username wajib diisi'", "", "[ ]"],
            ["1.7", "Login dengan password kosong saja", "1. Buka aplikasi\n2. Isi email/username\n3. Biarkan field password kosong\n4. Klik tombol 'MASUK'", "identifier: admin@mail.com\npassword: (kosong)", "Muncul pesan validasi 'Password wajib diisi'", "", "[ ]"],
            ["1.8", "Toggle visibilitas password", "1. Buka aplikasi\n2. Isi field password\n3. Klik ikon mata di field password", "password: password123", "Password berubah dari tersembunyi (*) menjadi terlihat (teks biasa), atau sebaliknya", "", "[ ]"],
            ["1.9", "Logout dari aplikasi", "1. Login ke aplikasi\n2. Navigasi ke menu logout\n3. Klik logout", "—", "Kembali ke halaman login, session dihapus", "", "[ ]"],
        ]
    },
    "2. Dashboard": {
        "desc": "Dashboard Admin, Manager, Driver",
        "tests": [
            ["2.1", "Dashboard Admin tampil sesuai role", "1. Login sebagai Admin\n2. Perhatikan tampilan dashboard", "Role: admin", "Dashboard admin tampil dengan menu: Kelola Perca, Kelola Penjahit, Kelola Majun, Kelola Expedisi", "", "[ ]"],
            ["2.2", "Dashboard Manager tampil sesuai role", "1. Login sebagai Manager\n2. Perhatikan tampilan dashboard", "Role: manager", "Dashboard manager tampil dengan menu manajemen: Kelola Partner, Kelola Pabrik, dan ringkasan data", "", "[ ]"],
            ["2.3", "Dashboard Driver tampil sesuai role", "1. Login sebagai Driver\n2. Perhatikan tampilan dashboard", "Role: driver", "Dashboard driver tampil dengan informasi pengiriman dan tugas", "", "[ ]"],
            ["2.4", "Profil user tampil di dashboard", "1. Login sebagai user\n2. Perhatikan card profil di dashboard", "User yang sudah login", "Card profil menampilkan nama, role, dan info user", "", "[ ]"],
            ["2.5", "Navigasi bottom bar berfungsi", "1. Login ke aplikasi\n2. Klik setiap item di bottom navigation bar", "—", "Setiap tab menampilkan halaman yang sesuai", "", "[ ]"],
        ]
    },
    "3. Kelola Perca": {
        "desc": "Input Stok Perca, Transaksi Ambil Perca, Riwayat",
        "tests": [
            ["3.1", "Input stok perca baru berhasil", "1. Buka menu Kelola Perca\n2. Klik 'Tambah Stok Perca'\n3. Pilih pabrik dari dropdown\n4. Pilih tanggal\n5. Isi jenis perca dan berat\n6. Ambil foto bukti\n7. Klik Simpan", "Pabrik: Pabrik A\nTanggal: 2026-04-23\nJenis: kaos\nBerat: 45 kg", "Data stok tersimpan, muncul notifikasi sukses, kode karung otomatis 'K-45'", "", "[ ]"],
            ["3.2", "Input stok perca jenis kain", "1. Buka menu Tambah Stok Perca\n2. Pilih jenis perca 'kain'\n3. Isi berat\n4. Simpan", "Jenis: kain\nBerat: 30 kg", "Kode karung otomatis 'B-30'", "", "[ ]"],
            ["3.3", "Input multiple stok perca sekaligus", "1. Buka menu Tambah Stok Perca\n2. Tambahkan beberapa item perca\n3. Klik Simpan semua", "Item 1: kaos 45kg\nItem 2: kain 30kg\nItem 3: kaos 25kg", "Semua item tersimpan ke database dalam satu batch", "", "[ ]"],
            ["3.4", "Transaksi ambil perca oleh penjahit", "1. Buka menu 'Transaksi Perca'\n2. Pilih penjahit dari dropdown\n3. Pilih kode karung yang tersedia\n4. Isi jumlah karung\n5. Klik 'Tambah ke Daftar'\n6. Klik 'Submit'", "Penjahit: Budi\nKode: K-45\nJumlah: 2", "Transaksi tercatat, stok penjahit bertambah, stok karung berkurang", "", "[ ]"],
            ["3.5", "Peringatan sisa perca penjahit > 5 Kg", "1. Pilih penjahit yang masih punya sisa perca > 5 Kg\n2. Coba tambahkan transaksi baru", "Penjahit: Budi (sisa perca 10 Kg)", "Muncul dialog peringatan sisa perca sebelum melanjutkan transaksi", "", "[ ]"],
            ["3.6", "Filter stok karung setelah ditambahkan ke daftar", "1. Tambahkan karung K-45 (stok 3) ke daftar sebanyak 2\n2. Perhatikan dropdown kode karung", "Kode: K-45, stok awal: 3, ditambahkan: 2", "Dropdown menunjukkan sisa stok K-45 = 1, bukan hilang dari daftar", "", "[ ]"],
            ["3.7", "Riwayat pengambilan perca tampil", "1. Buka menu 'Riwayat Perca'\n2. Perhatikan daftar riwayat", "—", "Menampilkan riwayat: Nama Pabrik, Tanggal, Total Berat, Jenis, Foto Bukti", "", "[ ]"],
            ["3.8", "Riwayat transaksi perca tampil", "1. Buka menu 'Riwayat Transaksi Perca'\n2. Perhatikan daftar", "—", "Menampilkan riwayat transaksi ambil perca oleh penjahit", "", "[ ]"],
        ]
    },
    "4. Kelola Penjahit": {
        "desc": "CRUD Penjahit, Detail, Upah, Efisiensi",
        "tests": [
            ["4.1", "Tambah penjahit baru berhasil", "1. Buka menu Kelola Penjahit\n2. Klik 'Tambah Penjahit'\n3. Isi nama, no HP, alamat\n4. (Opsional) Ambil foto\n5. Klik Simpan", "Nama: Siti\nNo HP: 081234567890\nAlamat: Jl. Raya 1", "Data penjahit baru tersimpan dan muncul di daftar", "", "[ ]"],
            ["4.2", "Tambah penjahit duplikat", "1. Klik 'Tambah Penjahit'\n2. Isi data dengan nama yang sudah ada\n3. Klik Simpan", "Nama: Siti (sudah ada)", "Muncul pesan error 'Data penjahit sudah terdaftar'", "", "[ ]"],
            ["4.3", "Edit data penjahit", "1. Buka daftar penjahit\n2. Pilih penjahit\n3. Klik Edit\n4. Ubah nama/no HP/alamat\n5. Klik Simpan", "Nama baru: Siti Aminah\nNo HP baru: 089876543210", "Data penjahit berhasil diupdate", "", "[ ]"],
            ["4.4", "Edit data penjahit — ganti foto", "1. Pilih penjahit\n2. Klik Edit\n3. Ganti foto baru\n4. Klik Simpan", "Foto baru dipilih dari galeri", "Foto lama dihapus dari storage, foto baru diupload dan tersimpan", "", "[ ]"],
            ["4.5", "Hapus data penjahit", "1. Buka daftar penjahit\n2. Pilih penjahit\n3. Klik Hapus\n4. Konfirmasi hapus", "ID penjahit yang dipilih", "Data penjahit dan foto terkait dihapus dari database dan storage", "", "[ ]"],
            ["4.6", "Cari penjahit berdasarkan nama", "1. Buka daftar penjahit\n2. Ketik nama di kolom pencarian", "Query: \"Budi\"", "Hanya penjahit dengan nama mengandung 'Budi' yang tampil", "", "[ ]"],
            ["4.7", "Cari penjahit — query kosong", "1. Buka daftar penjahit\n2. Kosongkan kolom pencarian", "Query: (kosong)", "Menampilkan semua data penjahit", "", "[ ]"],
            ["4.8", "Detail penjahit menampilkan statistik efisiensi", "1. Buka daftar penjahit\n2. Klik salah satu penjahit untuk melihat detail", "Penjahit: Budi", "Menampilkan: Total Perca Diambil, Total Majun Disetor, Reff, Prediksi Majun, Sisa Perca", "", "[ ]"],
            ["4.9", "Penarikan upah penjahit berhasil", "1. Buka detail penjahit\n2. Klik 'Tarik Upah'\n3. Isi nominal\n4. Pilih tanggal\n5. Klik Simpan", "Nominal: Rp500.000\nTanggal: 2026-04-23", "Penarikan tercatat, saldo penjahit berkurang Rp500.000", "", "[ ]"],
            ["4.10", "Riwayat setor majun per penjahit tampil", "1. Buka detail penjahit\n2. Lihat tab riwayat setor majun", "Penjahit: Budi", "Daftar riwayat setor majun dengan tanggal, berat, dan upah earned", "", "[ ]"],
            ["4.11", "Riwayat penarikan upah per penjahit tampil", "1. Buka detail penjahit\n2. Lihat tab riwayat tarik upah", "Penjahit: Budi", "Daftar riwayat penarikan dengan tanggal dan nominal", "", "[ ]"],
            ["4.12", "Tampilan saldo dan stok perca per penjahit", "1. Buka daftar penjahit\n2. Perhatikan info di setiap card", "—", "Setiap card menampilkan total_stock (sisa perca) dan balance (saldo upah) dengan format Rupiah", "", "[ ]"],
        ]
    },
    "5. Kelola Pabrik": {
        "desc": "CRUD Data Pabrik — factory_list_screen.dart",
        "tests": [
            ["5.1", "Tambah pabrik baru berhasil", "1. Buka menu Kelola Pabrik\n2. Klik 'Tambah Pabrik'\n3. Isi nama pabrik, alamat, no telp\n4. Klik Simpan", "Nama: Pabrik Jaya\nAlamat: Jl. Industri 5\nNo Telp: 021-1234567", "Data pabrik baru tersimpan dan muncul di daftar", "", "[ ]"],
            ["5.2", "Tambah pabrik duplikat", "1. Klik 'Tambah Pabrik'\n2. Isi nama pabrik yang sudah ada\n3. Klik Simpan", "Nama: Pabrik Jaya (sudah ada)", "Muncul pesan error duplicate", "", "[ ]"],
            ["5.3", "Edit data pabrik", "1. Pilih pabrik dari daftar\n2. Klik Edit\n3. Ubah data\n4. Klik Simpan", "Nama baru: Pabrik Jaya Makmur", "Data pabrik berhasil diupdate", "", "[ ]"],
            ["5.4", "Hapus pabrik tanpa relasi", "1. Pilih pabrik tanpa data terkait\n2. Klik Hapus\n3. Konfirmasi", "Pabrik tanpa data stok perca", "Pabrik berhasil dihapus", "", "[ ]"],
            ["5.5", "Hapus pabrik yang masih ada relasi", "1. Pilih pabrik yang masih punya data stok perca\n2. Klik Hapus", "Pabrik dengan data stok perca terkait", "Muncul pesan error: tidak bisa menghapus pabrik yang masih memiliki data terkait", "", "[ ]"],
            ["5.6", "Cari pabrik berdasarkan nama", "1. Ketik nama di kolom pencarian", "Query: \"Jaya\"", "Hanya pabrik dengan nama mengandung 'Jaya' yang tampil", "", "[ ]"],
        ]
    },
    "6. Kelola Majun": {
        "desc": "Setor Majun, Setor Limbah, Riwayat, Statistik",
        "tests": [
            ["6.1", "Setor majun berhasil", "1. Buka menu Kelola Majun\n2. Klik 'Setor Majun'\n3. Pilih penjahit\n4. Isi berat majun\n5. Ambil foto bukti\n6. Klik Simpan", "Penjahit: Budi\nBerat: 10.5 kg", "Transaksi tersimpan, earned_wage dihitung otomatis = berat × harga/kg, stok penjahit berkurang", "", "[ ]"],
            ["6.2", "Setor limbah berhasil", "1. Buka menu Kelola Majun\n2. Klik 'Setor Limbah'\n3. Pilih penjahit\n4. Isi berat limbah\n5. Klik Simpan", "Penjahit: Budi\nBerat: 5.0 kg", "Transaksi tersimpan, stok penjahit berkurang TANPA menambah upah", "", "[ ]"],
            ["6.3", "Riwayat setor majun tampil", "1. Buka menu 'Riwayat Setor Majun'\n2. Perhatikan daftar", "—", "Menampilkan daftar: Nama Penjahit, Tanggal, Berat, Upah, Foto Bukti", "", "[ ]"],
            ["6.4", "Riwayat setor limbah tampil", "1. Buka menu 'Riwayat Setor Limbah'\n2. Perhatikan daftar", "—", "Menampilkan daftar riwayat setor limbah", "", "[ ]"],
            ["6.5", "Lihat/ubah harga majun per kg", "1. Buka menu Kelola Majun\n2. Lihat harga per kg saat ini\n3. Ubah harga\n4. Simpan", "Harga baru: Rp6.000/kg", "Harga berhasil diupdate, setor majun selanjutnya menggunakan harga baru", "", "[ ]"],
            ["6.6", "Statistik bulanan setor majun tampil", "1. Buka dashboard/statistik\n2. Lihat chart bulanan", "—", "Grafik/chart menampilkan data setor majun per bulan", "", "[ ]"],
            ["6.7", "Format Rupiah tampil benar", "1. Perhatikan tampilan angka di seluruh aplikasi", "Angka: 100000", "Ditampilkan sebagai 'Rp100.000' (tanpa spasi setelah Rp, titik sebagai pemisah ribuan)", "", "[ ]"],
        ]
    },
    "7. Kelola Expedisi": {
        "desc": "Kirim Expedisi, Riwayat, Stok Gudang, Setting",
        "tests": [
            ["7.1", "Tambah expedisi baru berhasil", "1. Buka menu Kelola Expedisi\n2. Klik 'Tambah Expedisi'\n3. Pilih partner/driver\n4. Isi jumlah karung dan berat total\n5. Ambil foto bukti pengiriman\n6. Klik Simpan", "Partner: Driver A\nJumlah karung: 10\nBerat: 250 kg", "Expedisi tercatat, stok gudang berkurang, foto bukti tersimpan", "", "[ ]"],
            ["7.2", "Stok gudang tersedia ditampilkan", "1. Buka menu Tambah Expedisi\n2. Perhatikan info stok yang tersedia", "—", "Menampilkan stok gudang = Total setor majun - Total expedisi", "", "[ ]"],
            ["7.3", "Hapus data expedisi", "1. Buka riwayat expedisi\n2. Pilih expedisi\n3. Klik Hapus\n4. Konfirmasi", "ID expedisi yang dipilih", "Data expedisi dihapus dari database, foto dihapus dari storage", "", "[ ]"],
            ["7.4", "Riwayat expedisi tampil", "1. Buka menu 'Riwayat Expedisi'", "—", "Menampilkan daftar expedisi dengan tanggal, partner, berat, foto bukti", "", "[ ]"],
            ["7.5", "Kelola partner expedisi", "1. Buka menu 'Kelola Partner Expedisi'\n2. Lihat daftar partner", "—", "Menampilkan daftar partner expedisi yang terdaftar", "", "[ ]"],
            ["7.6", "Lihat/ubah berat standar per karung", "1. Buka setting expedisi\n2. Ubah berat per karung\n3. Simpan", "Berat baru: 30 kg/karung", "Setting tersimpan, perhitungan selanjutnya menggunakan berat baru", "", "[ ]"],
        ]
    },
    "8. Kelola Partner": {
        "desc": "Tambah/Edit/Hapus Admin dan Driver",
        "tests": [
            ["8.1", "Tambah akun admin baru berhasil", "1. Login sebagai Manager\n2. Buka menu Kelola Partner > Admin\n3. Klik 'Tambah Admin'\n4. Isi nama, email, no HP, password\n5. Klik Simpan", "Nama: Admin Baru\nEmail: admin2@mail.com\nNo HP: 081234567890\nPassword: Pass@123", "Akun admin baru berhasil dibuat via Edge Function, muncul di daftar admin", "", "[ ]"],
            ["8.2", "Tambah akun admin — email sudah terdaftar", "1. Klik 'Tambah Admin'\n2. Isi email yang sudah ada\n3. Klik Simpan", "Email: admin@mail.com (sudah ada)", "Muncul pesan error bahwa email sudah terdaftar", "", "[ ]"],
            ["8.3", "Tambah akun driver baru berhasil", "1. Login sebagai Manager\n2. Buka menu Kelola Partner > Driver\n3. Klik 'Tambah Driver'\n4. Isi nama, email, no HP, password\n5. Klik Simpan", "Nama: Driver Baru\nEmail: driver2@mail.com\nNo HP: 089876543210\nPassword: Pass@456", "Akun driver baru berhasil dibuat via Edge Function, muncul di daftar driver", "", "[ ]"],
            ["8.4", "Tambah akun driver — email sudah terdaftar", "1. Klik 'Tambah Driver'\n2. Isi email yang sudah ada\n3. Klik Simpan", "Email: driver@mail.com (sudah ada)", "Muncul pesan error bahwa email sudah terdaftar", "", "[ ]"],
            ["8.5", "Edit data admin", "1. Buka daftar admin\n2. Pilih admin\n3. Klik Edit\n4. Ubah nama/email/no HP\n5. Klik Simpan", "Nama baru: Admin Updated\nNo HP baru: 081111111111", "Data admin berhasil diupdate", "", "[ ]"],
            ["8.6", "Edit data admin — ubah password", "1. Pilih admin\n2. Klik Edit\n3. Isi field password baru\n4. Klik Simpan", "Password baru: NewPass@789", "Password admin berhasil diubah, bisa login dengan password baru", "", "[ ]"],
            ["8.7", "Edit data admin — tanpa ubah password", "1. Pilih admin\n2. Klik Edit\n3. Ubah nama saja, biarkan password kosong\n4. Klik Simpan", "Nama baru: Admin A\nPassword: (kosong)", "Hanya nama yang berubah, password tetap sama", "", "[ ]"],
            ["8.8", "Edit data driver", "1. Buka daftar driver\n2. Pilih driver\n3. Klik Edit\n4. Ubah data\n5. Klik Simpan", "Nama baru: Driver Updated\nAlamat: Jl. Baru 10", "Data driver berhasil diupdate termasuk alamat", "", "[ ]"],
            ["8.9", "Hapus akun admin", "1. Buka daftar admin\n2. Pilih admin\n3. Klik Hapus\n4. Konfirmasi", "Admin yang dipilih", "Akun admin dihapus dari auth.users dan profiles (cascade)", "", "[ ]"],
            ["8.10", "Hapus akun driver", "1. Buka daftar driver\n2. Pilih driver\n3. Klik Hapus\n4. Konfirmasi", "Driver yang dipilih", "Akun driver dihapus dari auth.users dan profiles (cascade)", "", "[ ]"],
            ["8.11", "Cari admin berdasarkan nama", "1. Buka daftar admin\n2. Ketik di kolom pencarian", "Query: \"Admin\"", "Hanya admin dengan nama mengandung 'Admin' yang tampil", "", "[ ]"],
            ["8.12", "Cari driver berdasarkan nama", "1. Buka daftar driver\n2. Ketik di kolom pencarian", "Query: \"Driver\"", "Hanya driver dengan nama mengandung 'Driver' yang tampil", "", "[ ]"],
            ["8.13", "Normalisasi nomor HP (format 08 → +62)", "1. Tambah/Edit staff\n2. Isi no HP dengan format 08xxx\n3. Simpan", "No HP: 081234567890", "Nomor disimpan dengan format yang dinormalisasi", "", "[ ]"],
            ["8.14", "Validasi form — field wajib kosong", "1. Klik Tambah Admin/Driver\n2. Biarkan semua field kosong\n3. Klik Simpan", "Semua field kosong", "Muncul pesan validasi untuk setiap field wajib", "", "[ ]"],
        ]
    },
    "9. Notifikasi": {
        "desc": "WhatsApp Notification — admin_notifications_screen.dart",
        "tests": [
            ["9.1", "Daftar notifikasi tampil", "1. Buka menu Notifikasi\n2. Lihat daftar notifikasi WhatsApp", "—", "Menampilkan daftar notifikasi yang sudah dikirim/antrian", "", "[ ]"],
            ["9.2", "Status notifikasi tampil benar", "1. Lihat notifikasi dalam daftar\n2. Perhatikan status setiap item", "—", "Setiap notifikasi menampilkan status: pending/sent/failed", "", "[ ]"],
        ]
    },
    "10. Upload & Kompresi": {
        "desc": "Upload Foto dan Kompresi Gambar",
        "tests": [
            ["10.1", "Upload foto bukti — gambar besar", "1. Pada fitur yang memerlukan foto (setor majun/perca)\n2. Ambil foto dengan kamera (resolusi tinggi)\n3. Simpan", "Foto 4032x3024px, 5MB", "Foto dikompres otomatis (max 1024px, quality 80), upload berhasil, URL tersimpan", "", "[ ]"],
            ["10.2", "Upload foto bukti — gambar kecil", "1. Pilih foto yang sudah kecil\n2. Simpan", "Foto 640x480px, 200KB", "Foto tidak diperbesar, upload berhasil", "", "[ ]"],
            ["10.3", "Upload foto gagal — tidak ada koneksi", "1. Matikan koneksi internet\n2. Coba upload foto", "Foto valid, tanpa internet", "Muncul pesan error 'Gagal upload gambar'", "", "[ ]"],
        ]
    },
}

def create_sheet(ws, module_name, desc, tests):
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = f"Modul {module_name} — {desc}"
    cell.font = MODULE_FONT
    cell.fill = MODULE_FILL
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, test in enumerate(tests, 3):
        for col_idx, value in enumerate(test, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = CENTER
            elif col_idx == 7:
                cell.alignment = CENTER
                cell.font = BOLD_FONT
            else:
                cell.alignment = WRAP
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A3"

# Create sheets
first = True
for module_name, data in MODULES.items():
    if first:
        ws = wb.active
        ws.title = module_name
        first = False
    else:
        ws = wb.create_sheet(title=module_name)
    create_sheet(ws, module_name, data["desc"], data["tests"])

# Summary sheet
ws_s = wb.create_sheet(title="Ringkasan", index=0)
ws_s.merge_cells("A1:D1")
c = ws_s["A1"]
c.value = "Ringkasan Blackbox Testing — Aplikasi Majunkita"
c.font = Font(bold=True, size=14, color="0D8043")
c.alignment = Alignment(vertical="center")

ws_s.merge_cells("A2:D2")
c = ws_s["A2"]
c.value = "Tanggal: 23 April 2026 | Teknik: Equivalence Partitioning, Boundary Value Analysis"
c.font = Font(size=10, color="666666")

sh = ["No", "Modul", "Deskripsi", "Jumlah Kasus Uji"]
for i, h in enumerate(sh, 1):
    c = ws_s.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
ws_s.column_dimensions["A"].width = 6
ws_s.column_dimensions["B"].width = 22
ws_s.column_dimensions["C"].width = 48
ws_s.column_dimensions["D"].width = 20

total = 0
for i, (name, data) in enumerate(MODULES.items(), 5):
    cnt = len(data["tests"]); total += cnt
    for col, val in enumerate([i-4, name, data["desc"], cnt], 1):
        c = ws_s.cell(row=i, column=col, value=val)
        c.font = NORMAL_FONT; c.border = THIN_BORDER
        c.alignment = CENTER if col in [1,4] else WRAP

tr = 5 + len(MODULES)
ws_s.cell(row=tr, column=3, value="TOTAL").font = BOLD_FONT
ws_s.cell(row=tr, column=3).alignment = Alignment(horizontal="right")
ws_s.cell(row=tr, column=3).border = THIN_BORDER
ws_s.cell(row=tr, column=4, value=total).font = Font(bold=True, size=11, color="0D8043")
ws_s.cell(row=tr, column=4).alignment = CENTER
ws_s.cell(row=tr, column=4).border = THIN_BORDER
ws_s.freeze_panes = "A5"

output = "/home/dwirez/PROJEK/T_A/majunkita/docs/blackbox_testing.xlsx"
wb.save(output)
print(f"✅ Saved: {output}")
print(f"📊 Total test cases: {total} across {len(MODULES)} modules")
